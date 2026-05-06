"""
Выводит все Telegram-каналы и группы текущего аккаунта,
включая те, от которых пользователь отписался (через Takeout API).
Поддержка вывода: stdout, .txt, .pdf (с кликабельными ссылками).
"""
import os
import sys
import asyncio
import argparse
from pathlib import Path
from datetime import datetime
from contextlib import contextmanager
from xml.sax.saxutils import escape as xml_escape

from telethon import TelegramClient
from telethon.tl.functions.channels import (
    GetLeftChannelsRequest,
    GetFullChannelRequest,
    GetParticipantRequest,
)
from telethon.tl.functions.account import (
    FinishTakeoutSessionRequest,
    InitTakeoutSessionRequest,
)
from telethon.tl.functions import InvokeWithTakeoutRequest
from telethon.tl.types import Channel, PeerChannel
from telethon.errors.rpcerrorlist import TakeoutInitDelayError

try:
    from wcwidth import wcswidth as _wcswidth
except ImportError:
    _wcswidth = None


# ----- Width helpers --------------------------------------------------------

def vwidth(s: str) -> int:
    """Визуальная ширина строки в терминальных колонках.

    Учитывает CJK-символы и эмодзи через ``wcwidth``, если он установлен.
    Иначе возвращает обычную длину строки (``len``) — приблизительно, но
    достаточно для ASCII/кириллицы.

    :param s: Произвольная строка.
    :return: Количество визуальных колонок, занимаемых строкой.
    """
    if _wcswidth is not None:
        w = _wcswidth(s)
        if w >= 0:
            return w
    return len(s)


def vpad(s: str, width: int, truncate: bool = True) -> str:
    """Добивает строку пробелами до заданной визуальной ширины.

    Если строка длиннее ``width`` и ``truncate=True``, обрезает её и
    дописывает символ многоточия ``…``. Если ``truncate=False`` — возвращает
    строку без изменений (без обрезки и без паддинга).

    :param s: Исходная строка.
    :param width: Целевая визуальная ширина в колонках.
    :param truncate: Разрешать ли обрезку слишком длинных строк.
    :return: Строка фиксированной визуальной ширины.
    """
    w = vwidth(s)
    if w <= width:
        return s + " " * (width - w)
    if not truncate:
        return s
    out = ""
    cur = 0
    for ch in s:
        cw = vwidth(ch)
        if cur + cw > width - 1:
            break
        out += ch
        cur += cw
    out += "…"
    cur += 1
    if cur < width:
        out += " " * (width - cur)
    return out


# ----- Config ---------------------------------------------------------------

def load_env(path: Path) -> None:
    """Загружает переменные окружения из ``.env``-файла.

    Простейший парсер: строки вида ``KEY=VALUE``, комментарии (``#``) и
    пустые строки игнорируются. Кавычки вокруг значения снимаются. Уже
    установленные в окружении переменные **не перезаписываются**
    (используется ``setdefault``).

    :param path: Путь к ``.env``-файлу. Если файла нет — функция тихо выходит.
    """
    if not path.exists():
        return
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


# ----- CLI ------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    """Разбирает аргументы командной строки.

    Конфигурирует ``argparse`` с описанием на русском, набором флагов
    (``--type``, ``--output``, ``--format``, ``--no-current``, ``--no-left``,
    ``--session``, ``--poll``, ``--lookup``) и примерами использования.

    :return: ``argparse.Namespace`` с заполненными полями.
    """
    p = argparse.ArgumentParser(
        prog="list_channels.py",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description=(
            "Выгружает список Telegram-каналов и групп текущего аккаунта.\n"
            "Включает активные подписки и покинутые каналы (через Takeout API).\n\n"
            "Требуется API_ID / API_HASH с https://my.telegram.org\n"
            "(задать в окружении или в файле .env рядом со скриптом)."
        ),
        epilog=(
            "Примеры:\n"
            "  python list_channels.py\n"
            "  python list_channels.py --type channels\n"
            "  python list_channels.py --type groups -o groups.txt\n"
            "  python list_channels.py -o report.pdf       # PDF с кликабельными ссылками\n"
            "  python list_channels.py --no-left\n"
            "  python list_channels.py --lookup 1312540285\n"
        ),
    )
    p.add_argument(
        "-t", "--type",
        dest="type",
        choices=("all", "channels", "groups"),
        default="all",
        help="Какие типы выводить (по умолчанию: all). "
             "'channels' = только broadcast-каналы; 'groups' = только супергруппы.",
    )
    p.add_argument(
        "-o", "--output",
        type=Path,
        default=None,
        help="Записать отчёт в файл вместо stdout. "
             "Формат определяется по расширению: .pdf → PDF (кликабельные ссылки), "
             "иначе — обычный текст.",
    )
    p.add_argument(
        "-f", "--format",
        choices=("auto", "txt", "pdf", "md", "xlsx"),
        default="auto",
        help="Формат вывода (по умолчанию: auto — по расширению файла: "
             ".pdf/.md/.xlsx/.txt).",
    )
    p.add_argument(
        "--no-current",
        action="store_true",
        help="Пропустить секцию активных подписок.",
    )
    p.add_argument(
        "--no-left",
        action="store_true",
        help="Пропустить покинутые каналы (без Takeout — не нужен аппрув).",
    )
    p.add_argument(
        "--session",
        default="anon",
        help="Имя Telethon-сессии (по умолчанию: anon).",
    )
    p.add_argument(
        "--poll",
        type=int,
        default=15,
        help="Интервал ретраев Takeout-инициации в секундах (по умолчанию: 15).",
    )
    p.add_argument(
        "--with-joined",
        action="store_true",
        help="Дозапросить дату вступления (joined_at) для активных подписок. "
             "Делает +1 RPC на канал — медленно при сотнях каналов.",
    )
    p.add_argument(
        "--with-activity",
        action="store_true",
        help="Дозапросить дату последнего сообщения (last_message_at). "
             "Делает +1 RPC на канал — медленно при сотнях каналов.",
    )
    p.add_argument(
        "--lookup",
        type=int,
        nargs="+",
        metavar="ID",
        help="Найти канал(ы) по id и показать подробности. "
             "В этом режиме секции активных/покинутых не выводятся.",
    )
    return p.parse_args()


def resolve_format(args) -> str:
    """Определяет фактический формат вывода.

    Если ``--format`` задан явно (не ``auto``) — возвращает его. Иначе
    угадывает формат по расширению ``--output`` (``.pdf`` / ``.md`` /
    ``.markdown`` / ``.xlsx``); если ничего не подошло — ``txt``.

    :param args: Результат ``parse_args``.
    :return: Один из ``"txt" | "pdf" | "md" | "xlsx"``.
    """
    if args.format != "auto":
        return args.format
    if args.output:
        ext = args.output.suffix.lower()
        if ext == ".pdf":
            return "pdf"
        if ext in (".md", ".markdown"):
            return "md"
        if ext == ".xlsx":
            return "xlsx"
    return "txt"


# ----- Record helpers -------------------------------------------------------

def type_of(e) -> str:
    """Возвращает тип сущности Telegram: ``channel`` или ``group``.

    Broadcast-канал в Telegram имеет флаг ``broadcast=True``; супергруппа —
    ``broadcast=False`` (в Telethon это всё ``Channel``-объекты).

    :param e: Telethon-сущность ``Channel``.
    :return: ``"channel"`` для broadcast-канала, иначе ``"group"``.
    """
    return "channel" if getattr(e, "broadcast", False) else "group"


def keep(entity, type_filter: str) -> bool:
    """Проверяет, проходит ли сущность через фильтр по типу.

    :param entity: Telethon-сущность ``Channel``.
    :param type_filter: ``"all"`` / ``"channels"`` / ``"groups"``.
    :return: ``True``, если сущность нужно оставить в выборке.
    """
    t = type_of(entity)
    if type_filter == "all":
        return True
    if type_filter == "channels":
        return t == "channel"
    if type_filter == "groups":
        return t == "group"
    return False


def _fmt_date(d) -> str:
    """Форматирует ``datetime`` как ``YYYY-MM-DD`` или возвращает ``""``.

    :param d: ``datetime`` или ``None``.
    :return: Строка даты или пустая строка.
    """
    if d is None:
        return ""
    try:
        return d.strftime("%Y-%m-%d")
    except Exception:
        return ""


def to_record(entity) -> dict:
    """Конвертирует Telethon-сущность в плоскую запись для отчёта.

    :param entity: Telethon-сущность ``Channel``.
    :return: Словарь с ключами ``id``, ``title``, ``username``, ``link``,
        ``type``, ``created_at``, ``members``, ``flags``, ``joined_at``,
        ``last_message_at``. ``flags`` — строка через запятую из набора
        ``verified``/``scam``/``fake``/``restricted``. ``joined_at`` и
        ``last_message_at`` заполняются только при флагах
        ``--with-joined``/``--with-activity``.
    """
    uname = getattr(entity, "username", None)
    flags = []
    if getattr(entity, "verified", False):
        flags.append("verified")
    if getattr(entity, "scam", False):
        flags.append("scam")
    if getattr(entity, "fake", False):
        flags.append("fake")
    if getattr(entity, "restricted", False):
        flags.append("restricted")
    return {
        "id": entity.id,
        "title": getattr(entity, "title", "") or "",
        "username": f"@{uname}" if uname else "-",
        "link": f"https://t.me/{uname}" if uname else "",
        "type": type_of(entity),
        "created_at": _fmt_date(getattr(entity, "date", None)),
        "members": getattr(entity, "participants_count", None),
        "flags": ",".join(flags),
        "joined_at": "",
        "last_message_at": "",
    }


# ----- Telegram fetchers ----------------------------------------------------

async def fetch_current(client, type_filter: str, log) -> list[dict]:
    """Загружает активные подписки пользователя.

    Обходит все диалоги (``iter_dialogs``), оставляет только ``Channel``
    (broadcast-каналы и супергруппы), применяет фильтр ``type_filter`` и
    дедуплицирует по ``id``.

    :param client: Авторизованный ``TelegramClient``.
    :param type_filter: ``"all"`` / ``"channels"`` / ``"groups"``.
    :param log: Колбэк для прогресс-логов (печатает в ``stderr``).
    :return: Список записей в формате ``to_record``.
    """
    out = []
    seen_ids: set[int] = set()
    async for d in client.iter_dialogs():
        e = d.entity
        if not isinstance(e, Channel):
            continue
        if e.id in seen_ids:
            continue
        if not keep(e, type_filter):
            continue
        seen_ids.add(e.id)
        out.append(to_record(e))
    log(f"  получено активных записей: {len(out)}")
    return out


async def fetch_left(client, type_filter: str, exclude_ids: set[int], poll: int, log) -> list[dict]:
    """Загружает покинутые каналы через Telegram Takeout API.

    Telegram отдаёт список ``GetLeftChannelsRequest`` только в рамках
    Takeout-сессии, которую пользователь должен подтвердить в официальном
    клиенте (всплывающее уведомление «Запрос на экспорт данных»). Метод:

    1. Сбрасывает старый ``takeout_id`` в сессии Telethon.
    2. В цикле инициирует ``InitTakeoutSessionRequest``, ловя
       ``TakeoutInitDelayError`` и засыпая на ``poll`` секунд между
       попытками — пока пользователь не подтвердит запрос.
    3. Постранично выгружает покинутые каналы через
       ``InvokeWithTakeoutRequest(GetLeftChannelsRequest)``.
    4. В ``finally`` корректно закрывает Takeout-сессию.

    :param client: Авторизованный ``TelegramClient``.
    :param type_filter: ``"all"`` / ``"channels"`` / ``"groups"``.
    :param exclude_ids: id, которые уже есть в активных подписках —
        исключаются, чтобы не было дублей в отчёте.
    :param poll: Интервал ретраев инициации Takeout, секунды.
    :param log: Колбэк для прогресс-логов.
    :return: Список записей покинутых каналов в формате ``to_record``.
    """
    log("  сбрасываю прежний takeout_id (если был)")
    client.session.takeout_id = None
    client.session.save()

    log("  инициирую takeout-сессию...")
    attempt = 0
    while True:
        try:
            init = await client(InitTakeoutSessionRequest(
                contacts=False,
                message_users=False,
                message_chats=False,
                message_megagroups=False,
                message_channels=False,
                files=False,
            ))
            break
        except TakeoutInitDelayError as ex:
            attempt += 1
            wait_total = getattr(ex, "seconds", 86400)
            log(
                f"  Telegram требует подтверждение. Открой офиц. клиент → "
                f"уведомление 'Запрос на экспорт данных' → Подтвердить. "
                f"Серверная задержка={wait_total}с. Попытка #{attempt} через {poll}с..."
            )
            await asyncio.sleep(poll)

    takeout_id = init.id
    log(f"  takeout_id={takeout_id}")

    out = []
    seen_ids: set[int] = set()
    try:
        offset = 0
        fetched = 0
        total = None
        while True:
            res = await client(InvokeWithTakeoutRequest(
                takeout_id, GetLeftChannelsRequest(offset=offset)
            ))
            chats = res.chats
            if not chats:
                break
            total = getattr(res, "count", len(chats))
            for c in chats:
                if c.id in exclude_ids or c.id in seen_ids:
                    continue
                if not keep(c, type_filter):
                    seen_ids.add(c.id)
                    continue
                seen_ids.add(c.id)
                out.append(to_record(c))
            fetched += len(chats)
            if total is not None and fetched >= total:
                break
            offset += len(chats)
    finally:
        try:
            await client(InvokeWithTakeoutRequest(
                takeout_id, FinishTakeoutSessionRequest(success=True)
            ))
        except Exception as ex:
            log(f"  предупреждение: закрытие takeout: {ex}")
        client.session.takeout_id = None
        client.session.save()

    log(f"  получено покинутых записей: {len(out)} (после фильтра)")
    return out


async def enrich_joined(client, items: list[dict], log) -> None:
    """Дозаполняет ``joined_at`` для каждой записи.

    Использует ``GetParticipantRequest(channel, 'me')`` →
    ``ChannelParticipantSelf.date``. Стоит +1 RPC на канал. При ошибке
    (нет доступа, флуд, удалён) поле остаётся пустым, инцидент логируется.

    :param client: Авторизованный ``TelegramClient``.
    :param items: Список записей ``to_record``; модифицируется in-place.
    :param log: Колбэк для прогресс-логов.
    """
    for it in items:
        try:
            ch = await client.get_input_entity(PeerChannel(it["id"]))
            res = await client(GetParticipantRequest(channel=ch, participant="me"))
            d = getattr(res.participant, "date", None)
            it["joined_at"] = _fmt_date(d)
        except Exception as ex:
            log(f"  предупреждение: joined id={it['id']}: {ex}")


async def enrich_activity(client, items: list[dict], log) -> None:
    """Дозаполняет ``last_message_at`` (дата последнего сообщения).

    Берёт первое сообщение через ``iter_messages(limit=1)``. Стоит +1 RPC
    на канал. Для канала без сообщений или при ошибке поле остаётся пустым.

    :param client: Авторизованный ``TelegramClient``.
    :param items: Список записей ``to_record``; модифицируется in-place.
    :param log: Колбэк для прогресс-логов.
    """
    for it in items:
        try:
            ch = await client.get_input_entity(PeerChannel(it["id"]))
            async for msg in client.iter_messages(ch, limit=1):
                if msg and msg.date:
                    it["last_message_at"] = _fmt_date(msg.date)
                break
        except Exception as ex:
            log(f"  предупреждение: activity id={it['id']}: {ex}")


async def lookup_records(
    client,
    ids: list[int],
    log,
    with_joined: bool = False,
    with_activity: bool = False,
) -> list[dict]:
    """Находит каналы по их числовым id и собирает подробности.

    Для каждого id вызывает ``get_entity(PeerChannel(id))``, затем
    ``GetFullChannelRequest`` для извлечения описания (``about``) и
    точного числа участников (``participants``). Дополнительно при
    флагах ``with_joined``/``with_activity`` дозаполняет ``joined_at``
    и ``last_message_at``. Если канал недоступен — запись помечается
    полем ``error`` с текстом исключения.

    :param client: Авторизованный ``TelegramClient``.
    :param ids: Список числовых id каналов.
    :param log: Колбэк для прогресс-логов.
    :param with_joined: Запрашивать дату вступления.
    :param with_activity: Запрашивать дату последнего сообщения.
    :return: Список словарей с ключами ``to_record`` + ``about``,
        ``participants``, ``error``.
    """
    out = []
    for cid in ids:
        log(f"  ищу id={cid}...")
        rec = {
            "id": cid, "title": "", "username": "-", "link": "",
            "type": "?", "created_at": "", "members": None, "flags": "",
            "joined_at": "", "last_message_at": "",
            "about": "", "participants": None, "error": None,
        }
        try:
            entity = await client.get_entity(PeerChannel(cid))
        except Exception as ex:
            rec["error"] = str(ex)
            out.append(rec)
            continue
        rec.update(to_record(entity))
        if not rec["link"]:
            rec["link"] = f"https://t.me/c/{entity.id}"
        try:
            full = await client(GetFullChannelRequest(entity))
            rec["about"] = (full.full_chat.about or "").strip()
            rec["participants"] = getattr(full.full_chat, "participants_count", None)
            if rec["members"] is None:
                rec["members"] = rec["participants"]
        except Exception as ex:
            log(f"  предупреждение: GetFullChannel id={cid}: {ex}")
        if with_joined:
            await enrich_joined(client, [rec], log)
        if with_activity:
            await enrich_activity(client, [rec], log)
        out.append(rec)
    return out


# ----- Text renderer --------------------------------------------------------

def banner_lines(args) -> list[str]:
    """Формирует строки шапки текстового отчёта.

    Использует псевдографические рамки (``╔═╗║╠╣╚╝``) и выравнивает
    содержимое через ``vpad``, чтобы рамка оставалась ровной даже при
    кириллических значениях.

    :param args: Результат ``parse_args`` (используются ``type``,
        ``no_current``, ``no_left``).
    :return: Список строк баннера, готовый к печати.
    """
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    width = 64
    L = []
    L.append("╔" + "═" * width + "╗")
    L.append("║" + "  Telegram: список каналов и групп".ljust(width) + "║")
    L.append("╠" + "═" * width + "╣")
    for label, value in [
        ("Время", now),
        ("Тип", args.type),
        ("Активные", "нет" if args.no_current else "да"),
        ("Покинутые", "нет" if args.no_left else "да (через Takeout)"),
    ]:
        line = f"║  {label:<10}: {value}"
        L.append(vpad(line, width + 1) + "║")
    L.append("╚" + "═" * width + "╝")
    return L


def _details_str(it: dict) -> str:
    """Собирает компактный «хвост» с метаданными канала.

    Включает только непустые поля: ``created``, ``members``, ``flags``,
    ``joined``, ``last``. Используется в текстовом и markdown-отчётах,
    чтобы не раздувать основные таблицы.

    :param it: Запись ``to_record``.
    :return: Строка вида ``created=… members=… flags=… joined=… last=…``.
    """
    parts = []
    if it.get("created_at"):
        parts.append(f"created={it['created_at']}")
    if it.get("members") is not None:
        parts.append(f"members={it['members']}")
    if it.get("flags"):
        parts.append(f"flags={it['flags']}")
    if it.get("joined_at"):
        parts.append(f"joined={it['joined_at']}")
    if it.get("last_message_at"):
        parts.append(f"last={it['last_message_at']}")
    return " ".join(parts)


def render_text_section(out, title: str, items: list[dict]) -> None:
    """Печатает секцию текстового отчёта (заголовок + таблица).

    Сортирует записи по ``(type, title.lower())``, динамически подбирает
    ширину колонок ``title`` и ``username`` под фактическое содержимое.
    После основной строки печатается отступленный «хвост» с метаданными
    (created/members/flags/joined/last), если хоть одно поле непустое.

    :param out: Файлоподобный объект для записи (``stdout`` или открытый файл).
    :param title: Заголовок секции (без счётчика — он добавится сам).
    :param items: Список записей в формате ``to_record``.
    """
    header = f"━━━ {title}  ({len(items)}) "
    print(header + "━" * max(0, 100 - vwidth(header)), file=out)
    if not items:
        print("  (пусто)", file=out)
        print(file=out)
        return

    title_w = min(max((vwidth(i["title"]) for i in items), default=10), 50)
    uname_w = max((vwidth(i["username"]) for i in items), default=10)
    items.sort(key=lambda x: (x["type"], x["title"].lower()))

    for it in items:
        link = it["link"] or "-"
        print(
            f"  [{it['type']:<7}] {vpad(it['title'], title_w)}  "
            f"{vpad(it['username'], uname_w)}  id={it['id']:<14}  {link}",
            file=out,
        )
        details = _details_str(it)
        if details:
            print(f"      {details}", file=out)
    print(file=out)


def render_text_lookup(out, items: list[dict]) -> None:
    """Печатает результат режима ``--lookup`` в текстовом виде.

    Каждая запись выводится блоком «ключ : значение». Если запись
    содержит поле ``error`` — печатает строку с ошибкой и пропускает
    остальные поля.

    :param out: Файлоподобный объект для записи.
    :param items: Результат ``lookup_records``.
    """
    print("━━━ Поиск по id ━━━".ljust(80, "━"), file=out)
    print(file=out)
    for it in items:
        if it.get("error"):
            print(f"  id={it['id']}: ОШИБКА — {it['error']}", file=out)
            print(file=out)
            continue
        print(f"  id          : {it['id']}", file=out)
        print(f"  тип         : {it['type']}", file=out)
        print(f"  название    : {it['title']}", file=out)
        print(f"  @username   : {it['username']}", file=out)
        print(f"  ссылка      : {it['link']}", file=out)
        if it.get("created_at"):
            print(f"  создан      : {it['created_at']}", file=out)
        if it.get("participants") is not None:
            print(f"  участников  : {it['participants']}", file=out)
        elif it.get("members") is not None:
            print(f"  участников  : {it['members']}", file=out)
        if it.get("flags"):
            print(f"  флаги       : {it['flags']}", file=out)
        if it.get("joined_at"):
            print(f"  вступил     : {it['joined_at']}", file=out)
        if it.get("last_message_at"):
            print(f"  посл. сообщ.: {it['last_message_at']}", file=out)
        if it.get("about"):
            print(f"  описание    : {it['about']}", file=out)
        print(file=out)


def write_text(path: Path | None, args, current, left, lookup) -> None:
    """Пишет отчёт в текстовом формате (TXT).

    Если ``path is None`` — пишет в ``sys.stdout``. Иначе открывает файл
    в UTF-8 на запись. В режиме ``--lookup`` печатается только результат
    поиска без секций активных/покинутых.

    :param path: Целевой путь либо ``None`` для stdout.
    :param args: Результат ``parse_args``.
    :param current: Записи активных подписок.
    :param left: Записи покинутых каналов.
    :param lookup: Результат ``--lookup`` либо ``None``.
    """
    @contextmanager
    def _open():
        if path is None:
            yield sys.stdout
        else:
            f = path.open("w", encoding="utf-8")
            try:
                yield f
            finally:
                f.close()

    with _open() as out:
        for line in banner_lines(args):
            print(line, file=out)
        print(file=out)

        if lookup is not None:
            render_text_lookup(out, lookup)
            return

        if not args.no_current:
            render_text_section(out, "Активные подписки", current)
        if not args.no_left:
            render_text_section(out, "Покинутые каналы (через takeout)", left)
        total_n = len(current) + len(left)
        print("━" * 100, file=out)
        print(f"  Всего записей: {total_n}", file=out)
        print("━" * 100, file=out)


# ----- PDF renderer ---------------------------------------------------------

def _register_mono_font() -> str:
    """Регистрирует моноширинный шрифт с поддержкой кириллицы для PDF.

    Перебирает кандидатов (Menlo на macOS, Courier New, DejaVu Sans Mono,
    Consolas) и регистрирует первый существующий под именем ``MonoCyr``.
    Если ничего не нашлось — возвращает встроенный ``Courier`` (кириллицу
    он не отрисует, но скрипт не упадёт).

    :return: Имя зарегистрированного шрифта для использования в reportlab.
    """
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    candidates = [
        ("/System/Library/Fonts/Menlo.ttc", 0),
        ("/System/Library/Fonts/Supplemental/Courier New.ttf", None),
        ("/Library/Fonts/Courier New.ttf", None),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSansMono.ttf", None),
        ("/usr/share/fonts/dejavu/DejaVuSansMono.ttf", None),
        ("C:/Windows/Fonts/consola.ttf", None),
        ("C:/Windows/Fonts/cour.ttf", None),
    ]
    for fpath, idx in candidates:
        if not Path(fpath).exists():
            continue
        try:
            kwargs = {"subfontIndex": idx} if idx is not None else {}
            pdfmetrics.registerFont(TTFont("MonoCyr", fpath, **kwargs))
            return "MonoCyr"
        except Exception:
            continue
    return "Courier"


def _esc(s: str) -> str:
    """XML-экранирование + замена пробелов на ``&nbsp;``.

    Используется внутри reportlab-параграфов, чтобы при переносе строк
    пробелы не схлопывались, а спецсимволы (``<``, ``>``, ``&``) не
    ломали разметку.

    :param s: Произвольный текст.
    :return: Безопасная для вставки в HTML-параграф reportlab строка.
    """
    return xml_escape(s).replace(" ", "&nbsp;")


def write_pdf(path: Path, args, current, left, lookup) -> None:
    """Пишет отчёт в PDF (альбомный A4, кликабельные ссылки).

    Использует ``reportlab.platypus`` для построения многостраничного
    документа с таблицами, чередованием цвета строк и гиперссылками в
    колонке ``link``. Шрифт подбирается в ``_register_mono_font`` — для
    корректной отрисовки кириллицы.

    :param path: Путь к выходному PDF (формат — landscape A4).
    :param args: Результат ``parse_args`` (используется для шапки и флагов).
    :param current: Записи активных подписок.
    :param left: Записи покинутых каналов.
    :param lookup: Результат ``--lookup`` либо ``None``.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors

    font_name = _register_mono_font()

    body = ParagraphStyle("body", fontName=font_name, fontSize=8, leading=10)
    body_link = ParagraphStyle("body_link", fontName=font_name, fontSize=8, leading=10,
                               textColor=colors.HexColor("#1a73e8"))
    head = ParagraphStyle("head", fontName=font_name, fontSize=12, leading=15,
                          spaceBefore=8, spaceAfter=4, textColor=colors.HexColor("#222222"))
    title = ParagraphStyle("title", fontName=font_name, fontSize=14, leading=18,
                           spaceAfter=8, textColor=colors.black)
    meta = ParagraphStyle("meta", fontName=font_name, fontSize=9, leading=11)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title="Telegram channels report",
    )
    story = []
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    story.append(Paragraph("Telegram: список каналов и групп", title))
    story.append(Paragraph(
        f"Время: {now} &nbsp;|&nbsp; Тип: {args.type} &nbsp;|&nbsp; "
        f"Активные: {'нет' if args.no_current else 'да'} &nbsp;|&nbsp; "
        f"Покинутые: {'нет' if args.no_left else 'да (Takeout)'}",
        meta,
    ))
    story.append(Spacer(1, 6))

    page_w = landscape(A4)[0] - 20 * mm
    col_widths = [
        16 * mm,   # type
        70 * mm,   # title
        34 * mm,   # username
        24 * mm,   # id
        55 * mm,   # details (created/members/flags/joined/last)
        page_w - (16 + 70 + 34 + 24 + 55) * mm,  # link (rest)
    ]

    table_style = TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("LEADING", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#f6f8fa")]),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e0e4ea")),
        ("FONTNAME", (0, 0), (-1, 0), font_name),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#222222")),
    ])

    def make_row(it: dict) -> list:
        """Собирает одну строку таблицы из записи ``to_record``.

        :param it: Запись канала/группы.
        :return: Список ``Paragraph``-ов в порядке колонок таблицы.
        """
        link_para = (
            Paragraph(f'<a href="{xml_escape(it["link"])}">{xml_escape(it["link"])}</a>',
                      body_link)
            if it["link"] else Paragraph("-", body)
        )
        details_lines = []
        if it.get("created_at"):
            details_lines.append(f"created: {it['created_at']}")
        if it.get("members") is not None:
            details_lines.append(f"members: {it['members']}")
        if it.get("flags"):
            details_lines.append(f"flags: {it['flags']}")
        if it.get("joined_at"):
            details_lines.append(f"joined: {it['joined_at']}")
        if it.get("last_message_at"):
            details_lines.append(f"last: {it['last_message_at']}")
        details_html = "<br/>".join(_esc(s) for s in details_lines) or "-"
        return [
            Paragraph(xml_escape(f"[{it['type']}]"), body),
            Paragraph(xml_escape(it["title"]), body),
            Paragraph(xml_escape(it["username"]), body),
            Paragraph(f"id={it['id']}", body),
            Paragraph(details_html, body),
            link_para,
        ]

    def section(title_text: str, items: list[dict]):
        """Добавляет в PDF-историю секцию: заголовок и таблицу записей.

        :param title_text: Заголовок секции (без счётчика).
        :param items: Список записей. Сортируется по типу и заголовку.
        """
        story.append(Paragraph(xml_escape(f"━━━ {title_text}  ({len(items)})"), head))
        if not items:
            story.append(Paragraph("(пусто)", body))
            return
        items.sort(key=lambda x: (x["type"], x["title"].lower()))
        rows = [["тип", "название", "@username", "id", "детали", "ссылка"]]
        for it in items:
            rows.append(make_row(it))
        t = Table(rows, colWidths=col_widths, repeatRows=1)
        t.setStyle(table_style)
        story.append(t)

    if lookup is not None:
        story.append(Paragraph("━━━ Поиск по id ━━━", head))
        for it in lookup:
            if it.get("error"):
                story.append(Paragraph(_esc(f"id={it['id']}: ОШИБКА — {it['error']}"), body))
                story.append(Spacer(1, 4))
                continue
            link_html = (
                f'<a href="{xml_escape(it["link"])}" color="#1a73e8">{_esc(it["link"])}</a>'
                if it["link"] else "-"
            )
            block = [
                f"id&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:&nbsp;{it['id']}",
                f"тип&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:&nbsp;{_esc(it['type'])}",
                f"название&nbsp;&nbsp;&nbsp;&nbsp;:&nbsp;{_esc(it['title'])}",
                f"@username&nbsp;&nbsp;&nbsp;:&nbsp;{_esc(it['username'])}",
                f"ссылка&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:&nbsp;{link_html}",
            ]
            if it.get("created_at"):
                block.append(f"создан&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:&nbsp;{_esc(it['created_at'])}")
            participants = it.get("participants")
            if participants is None:
                participants = it.get("members")
            if participants is not None:
                block.append(f"участников&nbsp;&nbsp;:&nbsp;{participants}")
            if it.get("flags"):
                block.append(f"флаги&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:&nbsp;{_esc(it['flags'])}")
            if it.get("joined_at"):
                block.append(f"вступил&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:&nbsp;{_esc(it['joined_at'])}")
            if it.get("last_message_at"):
                block.append(f"посл.&nbsp;сообщ.:&nbsp;{_esc(it['last_message_at'])}")
            if it.get("about"):
                block.append(f"описание&nbsp;&nbsp;&nbsp;&nbsp;:&nbsp;{_esc(it['about'])}")
            for line in block:
                story.append(Paragraph(line, body))
            story.append(Spacer(1, 6))
    else:
        if not args.no_current:
            section("Активные подписки", current)
        if not args.no_left:
            section("Покинутые каналы (через takeout)", left)
        total_n = len(current) + len(left)
        story.append(Spacer(1, 8))
        story.append(Paragraph(f"Всего записей: {total_n}", head))

    doc.build(story)


# ----- Markdown renderer ----------------------------------------------------

def _md_escape(s: str) -> str:
    """Экранирует значение для ячейки Markdown-таблицы.

    Заменяет ``|`` на ``\\|`` (иначе ломается разметка таблицы) и
    переводит переносы строк в пробелы.

    :param s: Произвольная строка (или ``None``).
    :return: Безопасное содержимое одной ячейки таблицы.
    """
    return (s or "").replace("|", "\\|").replace("\n", " ").replace("\r", " ")


def _md_link(it: dict) -> str:
    """Формирует Markdown-ссылку для записи или ``—``, если её нет.

    :param it: Запись ``to_record``.
    :return: Строка ``[url](url)`` либо ``"—"``.
    """
    if it["link"]:
        return f"[{_md_escape(it['link'])}]({it['link']})"
    return "—"


def _md_section(out, title: str, items: list[dict]) -> None:
    """Печатает одну секцию Markdown-отчёта.

    Заголовок второго уровня + таблица с пятью колонками. Записи
    сортируются по ``(type, title.lower())``.

    :param out: Файлоподобный объект для записи.
    :param title: Заголовок секции.
    :param items: Список записей ``to_record``.
    """
    print(f"## {title}  ({len(items)})\n", file=out)
    if not items:
        print("_(пусто)_\n", file=out)
        return
    has_joined = any(i.get("joined_at") for i in items)
    has_activity = any(i.get("last_message_at") for i in items)
    headers = ["тип", "название", "@username", "id", "создан", "участников", "флаги"]
    if has_joined:
        headers.append("вступил")
    if has_activity:
        headers.append("посл. сообщ.")
    headers.append("ссылка")
    print("| " + " | ".join(headers) + " |", file=out)
    print("|" + "|".join(["-----"] * len(headers)) + "|", file=out)
    items.sort(key=lambda x: (x["type"], x["title"].lower()))
    for it in items:
        members = it.get("members")
        members_s = str(members) if members is not None else "—"
        flags_s = it.get("flags") or "—"
        cells = [
            it["type"],
            _md_escape(it["title"]),
            f"`{_md_escape(it['username'])}`",
            f"`{it['id']}`",
            it.get("created_at") or "—",
            members_s,
            flags_s,
        ]
        if has_joined:
            cells.append(it.get("joined_at") or "—")
        if has_activity:
            cells.append(it.get("last_message_at") or "—")
        cells.append(_md_link(it))
        print("| " + " | ".join(cells) + " |", file=out)
    print(file=out)


def write_md(path: Path, args, current, left, lookup) -> None:
    """Пишет отчёт в Markdown.

    Шапка с метаданными, далее — секции активных и/или покинутых каналов
    либо одна таблица результатов ``--lookup``. В конце — итоговая
    строка «Всего записей».

    :param path: Целевой ``.md``-файл (UTF-8).
    :param args: Результат ``parse_args``.
    :param current: Записи активных подписок.
    :param left: Записи покинутых каналов.
    :param lookup: Результат ``--lookup`` либо ``None``.
    """
    with path.open("w", encoding="utf-8") as out:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print("# Telegram: список каналов и групп\n", file=out)
        print(
            f"**Время:** {now}  \n"
            f"**Тип:** `{args.type}`  \n"
            f"**Активные:** {'нет' if args.no_current else 'да'}  \n"
            f"**Покинутые:** {'нет' if args.no_left else 'да (Takeout)'}\n",
            file=out,
        )
        if lookup is not None:
            print("## Поиск по id\n", file=out)
            headers = [
                "id", "тип", "название", "@username",
                "создан", "участников", "флаги",
                "вступил", "посл. сообщ.", "ссылка", "описание",
            ]
            print("| " + " | ".join(headers) + " |", file=out)
            print("|" + "|".join(["----"] * len(headers)) + "|", file=out)
            for it in lookup:
                if it.get("error"):
                    print(
                        f"| `{it['id']}` | — | _ОШИБКА_ | — | — | — | — | — | — | — | "
                        f"{_md_escape(it['error'])} |",
                        file=out,
                    )
                    continue
                participants = it.get("participants")
                if participants is None:
                    participants = it.get("members")
                cells = [
                    f"`{it['id']}`",
                    it["type"],
                    _md_escape(it["title"]),
                    f"`{_md_escape(it['username'])}`",
                    it.get("created_at") or "—",
                    str(participants) if participants is not None else "—",
                    it.get("flags") or "—",
                    it.get("joined_at") or "—",
                    it.get("last_message_at") or "—",
                    _md_link(it),
                    _md_escape(it.get("about", "")),
                ]
                print("| " + " | ".join(cells) + " |", file=out)
            print(file=out)
            return
        if not args.no_current:
            _md_section(out, "Активные подписки", current)
        if not args.no_left:
            _md_section(out, "Покинутые каналы (через takeout)", left)
        print(f"---\n**Всего записей:** {len(current) + len(left)}", file=out)


# ----- XLSX renderer --------------------------------------------------------

def _xlsx_autofit(ws, max_widths: dict[int, int]) -> None:
    """Подгоняет ширину колонок XLSX под максимальную длину содержимого.

    Ширина зажимается в диапазоне ``[8, 80]`` плюс небольшой запас (``+2``),
    чтобы не было ни «слипшихся», ни чрезмерно растянутых колонок.

    :param ws: Лист openpyxl.
    :param max_widths: Словарь ``номер_колонки → максимальная длина строки``.
    """
    from openpyxl.utils import get_column_letter
    for col_idx, width in max_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 8), 80)


def _xlsx_write_table(ws, headers: list[str], rows: list[list], link_col: int | None) -> None:
    """Заполняет лист таблицей: шапка, строки, стили, гиперссылки.

    Шапка — синий фон с белым текстом, чётные строки — лёгкая «зебра»,
    тонкие серые границы у всех ячеек. Если задан ``link_col`` —
    значения этой колонки, начинающиеся с ``http``, превращаются в
    кликабельные гиперссылки. Первая строка фиксируется (``freeze_panes``).

    :param ws: Лист openpyxl.
    :param headers: Заголовки колонок.
    :param rows: Двумерный список значений (без шапки).
    :param link_col: Номер колонки с URL (1-based) либо ``None``.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_font = Font(name="Menlo", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    zebra_fill = PatternFill("solid", fgColor="F2F2F2")
    link_font = Font(name="Menlo", color="1A73E8", underline="single", size=10)
    body_font = Font(name="Menlo", size=10)
    center = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_widths: dict[int, int] = {}

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        max_widths[col_idx] = len(header)

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = body_font
            cell.alignment = center
            cell.border = border
            if r_idx % 2 == 0:
                cell.fill = zebra_fill
            sval = "" if value is None else str(value)
            max_widths[c_idx] = max(max_widths[c_idx], len(sval))
            if link_col is not None and c_idx == link_col and value and str(value).startswith("http"):
                cell.hyperlink = value
                cell.font = link_font

    ws.freeze_panes = "A2"
    _xlsx_autofit(ws, max_widths)


def write_xlsx(path: Path, args, current, left, lookup) -> None:
    """Пишет отчёт в Excel-книгу (.xlsx).

    Структура книги:

    - ``Сводка`` — метаданные и итоговые счётчики.
    - ``Активные`` — таблица активных подписок (если не ``--no-current``).
    - ``Покинутые`` — таблица покинутых каналов (если не ``--no-left``).
    - ``Поиск`` — результаты ``--lookup`` (вместо двух предыдущих листов).

    Используется openpyxl. Ссылки в колонке ``ссылка`` — кликабельные.

    :param path: Целевой ``.xlsx``-файл.
    :param args: Результат ``parse_args``.
    :param current: Записи активных подписок.
    :param left: Записи покинутых каналов.
    :param lookup: Результат ``--lookup`` либо ``None``.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    wb.remove(wb.active)

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    info = wb.create_sheet("Сводка")
    info["A1"] = "Telegram: список каналов и групп"
    info["A1"].font = Font(bold=True, size=14)
    info["A3"] = "Время"; info["B3"] = now
    info["A4"] = "Тип"; info["B4"] = args.type
    info["A5"] = "Активные"; info["B5"] = "нет" if args.no_current else "да"
    info["A6"] = "Покинутые"; info["B6"] = "нет" if args.no_left else "да (Takeout)"
    info["A7"] = "Активных записей"; info["B7"] = len(current)
    info["A8"] = "Покинутых записей"; info["B8"] = len(left)
    info["A9"] = "Всего"; info["B9"] = len(current) + len(left)
    for r in range(3, 10):
        info.cell(row=r, column=1).font = Font(bold=True)
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 40

    def section_sheet(name: str, items: list[dict]):
        """Создаёт отдельный лист с одной таблицей записей.

        :param name: Имя листа в книге Excel.
        :param items: Список записей ``to_record``.
        """
        ws = wb.create_sheet(name)
        items.sort(key=lambda x: (x["type"], x["title"].lower()))
        rows = [
            [
                it["type"], it["title"], it["username"], it["id"],
                it.get("created_at", ""),
                it.get("members") if it.get("members") is not None else "",
                it.get("flags", ""),
                it.get("joined_at", ""),
                it.get("last_message_at", ""),
                it["link"] or "",
            ]
            for it in items
        ]
        _xlsx_write_table(
            ws,
            ["тип", "название", "@username", "id",
             "создан", "участников", "флаги",
             "вступил", "посл. сообщ.", "ссылка"],
            rows,
            link_col=10,
        )

    if lookup is not None:
        ws = wb.create_sheet("Поиск")
        rows = []
        for it in lookup:
            if it.get("error"):
                rows.append([
                    it["id"], "—", f"ОШИБКА: {it['error']}", "—",
                    "—", "—", "—", "—", "—", "—", "—",
                ])
                continue
            participants = it.get("participants")
            if participants is None:
                participants = it.get("members")
            rows.append([
                it["id"], it["type"], it["title"], it["username"],
                it.get("created_at", ""),
                participants if participants is not None else "",
                it.get("flags", ""),
                it.get("joined_at", ""),
                it.get("last_message_at", ""),
                it["link"] or "",
                it.get("about", ""),
            ])
        _xlsx_write_table(
            ws,
            ["id", "тип", "название", "@username",
             "создан", "участников", "флаги",
             "вступил", "посл. сообщ.", "ссылка", "описание"],
            rows,
            link_col=10,
        )
    else:
        if not args.no_current:
            section_sheet("Активные", current)
        if not args.no_left:
            section_sheet("Покинутые", left)

    wb.save(str(path))


# ----- Main -----------------------------------------------------------------

async def run(args) -> None:
    """Главный сценарий: подключение, загрузка данных, запись отчёта.

    Шаги:

    1. Определяет формат вывода и валидирует наличие ``--output`` для
       бинарных форматов (PDF/MD/XLSX).
    2. Поднимает ``TelegramClient`` с указанной сессией. При первом
       запуске Telethon сам спросит телефон, код и пароль 2FA.
    3. В зависимости от флагов: либо ``--lookup``, либо обычный режим с
       загрузкой активных и/или покинутых каналов.
    4. Гарантированно отключается от Telegram в ``finally``.
    5. Делегирует запись соответствующему ``write_*``.

    :param args: Результат ``parse_args``.
    """
    fmt = resolve_format(args)
    if fmt in ("pdf", "md", "xlsx") and args.output is None:
        print(f"ОШИБКА: для формата {fmt} нужен --output PATH.", file=sys.stderr)
        sys.exit(2)

    def log(msg: str) -> None:
        """Печатает прогресс-сообщение в stderr с принудительным flush.

        Stderr используется, чтобы не смешиваться с отчётом, который
        может уходить в stdout.

        :param msg: Текст сообщения.
        """
        print(msg, file=sys.stderr, flush=True)

    log(f"→ Подключаюсь, сессия={args.session!r}...")
    client = TelegramClient(args.session, API_ID, API_HASH)
    await client.start()
    me = await client.get_me()
    log(f"→ Авторизован: {me.first_name} (@{me.username}) id={me.id}")

    current: list[dict] = []
    left: list[dict] = []
    lookup: list[dict] | None = None

    try:
        if args.lookup:
            log("→ Режим lookup: ищу по id...")
            lookup = await lookup_records(
                client, args.lookup, log,
                with_joined=args.with_joined,
                with_activity=args.with_activity,
            )
        else:
            if not args.no_current:
                log("→ Загружаю АКТИВНЫЕ подписки...")
                current = await fetch_current(client, args.type, log)
                if args.with_joined:
                    log("→ Дозапрашиваю даты вступления...")
                    await enrich_joined(client, current, log)
                if args.with_activity:
                    log("→ Дозапрашиваю активность (последние сообщения)...")
                    await enrich_activity(client, current, log)
            if not args.no_left:
                log("→ Загружаю ПОКИНУТЫЕ каналы через Takeout...")
                exclude = {r["id"] for r in current}
                left = await fetch_left(client, args.type, exclude, args.poll, log)
    finally:
        await client.disconnect()
        log("→ Отключён.")

    log(f"→ Запись отчёта (формат={fmt})...")
    if fmt == "pdf":
        write_pdf(args.output, args, current, left, lookup)
    elif fmt == "md":
        write_md(args.output, args, current, left, lookup)
    elif fmt == "xlsx":
        write_xlsx(args.output, args, current, left, lookup)
    else:
        write_text(args.output, args, current, left, lookup)
    if args.output:
        log(f"→ Готово: {args.output}")


def main() -> None:
    """Точка входа CLI.

    Разбирает аргументы, подгружает ``.env`` рядом со скриптом, проверяет
    наличие ``API_ID`` / ``API_HASH``, выставляет глобальные переменные
    для Telethon и запускает асинхронный ``run`` через ``asyncio.run``.

    Завершается с кодом ``2``, если переменные окружения не заданы.
    """
    args = parse_args()
    load_env(Path(__file__).parent / ".env")
    if "API_ID" not in os.environ or "API_HASH" not in os.environ:
        print(
            "ОШИБКА: API_ID / API_HASH не заданы. "
            "Положи их в .env рядом со скриптом или экспортируй в окружение.",
            file=sys.stderr,
        )
        sys.exit(2)

    global API_ID, API_HASH
    API_ID = int(os.environ["API_ID"])
    API_HASH = os.environ["API_HASH"]

    asyncio.run(run(args))


if __name__ == "__main__":
    main()
